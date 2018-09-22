'''
Standart cells:
A2, A3... - number of patient (1, 2, 3...)
B2, B3... - visit date (dd.mm.yyyy)
C2, C3... - type of visit (1, 2 or 3)
D2, D3... - consultation type (1, 2 or 3)
E2, E3... - full name of the patient
F2, F3... - gender (male or female)
G2, G3... - date of birth (dd.mm.yyyy)
H2, H3... - phone number
I2, I3... - ICPC-2-E code
J2, J3... - MKB-10 code
K2, K3... - ICPC-2-E process code
L2, L3... - category (1-11)
'''

'''
ABSTRACTION
'''
class Patient:
    def __init__(self, number, visit_date, visit_type, consultation_type, name,
                 gender, birthday, phone, ICPC, MKB10, ICPC_process, category):
        self.number = str(number)
        self.visit_date = str(visit_date)
        self.visit_type = str(visit_type)
        self.consultation_type = str(consultation_type)
        self.name = str(name)
        self.gender = 'm' if gender == 'чол.' else 'f'
        self.birthday = str(birthday)
        self.phone = str(phone)
        self.ICPC = str(ICPC)
        self.MKB10 = str(MKB10)
        self.ICPC_process = str(ICPC_process)
        self.category = str(category)

    def __repr__(self):
        return self.name

'''
FILTERS
'''
def check_visit(visit, patient):
    if patient.visit_type == visit:
        return patient

def check_consultation(consultation, patient):
    if patient.consultation_type == consultation:
        return patient

def check_gender(gender, patient):
    if patient.gender == gender:
        return patient

def check_ICPC(ICPC, patient):
    ICPC = ICPC.split()
    res = True
    for i in ICPC:
        if i not in patient.ICPC:
            res = False
    if res:
        return patient

def check_MKB10(MKB10, patient):
    MKB10 = MKB10.split()
    res = True
    for i in MKB10:
        if i not in patient.MKB10:
            res = False
    if res:
        return patient

def check_process(process, patient):
    process = process.split()
    res = True
    for i in process:
        if i not in patient.ICPC_process:
            res = False
    if res:
        return patient

def check_category(category, patient):
    if patient.category == category:
        return patient

'''
IMPORT PATIENTS INFO
'''
import openpyxl

patients = []

main_wb = openpyxl.load_workbook('main.xlsx')
sheet = main_wb['Sheet1']
max_patient = sheet.max_row

for i in range(max_patient-2):
    patient = Patient(
        sheet.cell(row=i+3, column=1).value,
        sheet.cell(row=i+3, column=2).value,
        sheet.cell(row=i+3, column=3).value,
        sheet.cell(row=i+3, column=4).value,
        sheet.cell(row=i+3, column=5).value,
        sheet.cell(row=i+3, column=6).value,
        sheet.cell(row=i+3, column=7).value,
        sheet.cell(row=i+3, column=8).value,
        sheet.cell(row=i+3, column=9).value,
        sheet.cell(row=i+3, column=10).value,
        sheet.cell(row=i+3, column=11).value,
        sheet.cell(row=i+3, column=12).value)
    patients.append(patient)

'''
WORKING LOOP
'''
command = ''

while command != 'exit':

    command = input('Enter the command from the list:' + 
'''
visit_type=n
consultation_type=n
gender=g
ICPC=all codes
MKB10=all codes
process=all codes
category=n
exit
''')

    answer = []

    if 'visit_type' in command:
        for p in patients:
            answer.append(check_visit(command[-1], p))
        while None in answer:
            answer.remove(None)

    elif 'consultation_type' in command:
        for p in patients:
            answer.append(check_consultation(command[-1], p))
        while None in answer:
            answer.remove(None)

    elif 'gender' in command:
        for p in patients:
            answer.append(check_gender(command.split('=')[1], p))
        while None in answer:
            answer.remove(None)

    elif 'ICPC' in command:
        for p in patients:
            answer.append(check_ICPC(command.split('=')[1], p))
        while None in answer:
            answer.remove(None)

    elif 'MKB10' in command:
        for p in patients:
            answer.append(check_MKB10(command.split('=')[1], p))
        while None in answer:
            answer.remove(None)

    elif 'process' in command:
        for p in patients:
            answer.append(check_process(command.split('=')[1], p))
        while None in answer:
            answer.remove(None)

    elif 'category' in command:
        for p in patients:
            answer.append(check_category(command.split('=')[1], p))
        while None in answer:
            answer.remove(None)

    print('\nThe amount of patients =', len(answer), '\n')