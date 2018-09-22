'''
Standart cells:
A2, A3... - number of patient (1, 2, 3...)
B2, B3... - visit date (dd.mm.yyyy)
C2, C3... - type of visit (1, 2 or 3)
D2, D3... - consultation type (1, 2, or 3)
E2, E3... - full name of the patient
F2, F3... - gender (male or female)
G2, G3... - date of birth (dd.mm.yyyy)
H2, H3... - phone number
I2, I3... - ICPC-2-E code
J2, J3... - MKB-10 code
K2, K3... - ICPC-2-E process code
L2, L3... - category (1-11)
'''

import openpyxl

main_wb = openpyxl.load_workbook('main.xlsx')
main_sheet = main_wb['Sheet1']

wb = openpyxl.load_workbook('doctor_1.xlsx')
sheet = wb['Sheet1']
max_patient = sheet.max_row
cells = [chr(i) for i in range(65, 65+12)]
for i in range(max_patient):
    patient_info = []
    for j in range(len(cells)):
        info = sheet.cell(row=i+1, column=j+1).value
        patient_info.append(info)
        main_sheet[cells[j]+str(i+1)] = str(info)

main_wb.create_sheet(title='Sheet2')
main_sheet = main_wb['Sheet2']

wb = openpyxl.load_workbook('doctor_2.xlsx')
sheet = wb['Sheet1']
max_patient = sheet.max_row
cells = [chr(i) for i in range(65, 65+12)]
for i in range(max_patient):
    patient_info = []
    for j in range(len(cells)):
        info = sheet.cell(row=i+1, column=j+1).value
        patient_info.append(info)
        main_sheet[cells[j]+str(i+1)] = str(info)

main_wb.save('main.xlsx')